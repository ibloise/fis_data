from __future__ import annotations

import json

from click.testing import CliRunner
from sqlalchemy import create_engine, text

from fis_data.cli import _expand_input_paths, cli


def _write_workbook(path, *, sheets: dict[str, list[list[object]]]) -> None:
    from openpyxl import Workbook

    workbook = Workbook()
    default = workbook.active
    workbook.remove(default)
    for sheet_name, rows in sheets.items():
        worksheet = workbook.create_sheet(sheet_name)
        for row in rows:
            worksheet.append(row)
    workbook.save(path)


def test_expand_input_paths_accepts_directories(tmp_path) -> None:
    first = tmp_path / "first.txt"
    second = tmp_path / "second.csv"
    nested_dir = tmp_path / "nested"
    nested_dir.mkdir()
    nested = nested_dir / "nested.txt"
    first.write_text("a", encoding="utf-8")
    second.write_text("b", encoding="utf-8")
    nested.write_text("c", encoding="utf-8")

    paths = _expand_input_paths((str(tmp_path),))

    assert paths == [first, second]


def test_expand_input_paths_accepts_recursive_excel_directories(tmp_path) -> None:
    flat = tmp_path / "farmacia.xlsx"
    nested_dir = tmp_path / "PCR_RUN"
    nested_dir.mkdir()
    nested = nested_dir / "PCR_RUN - Quantification Cq Results.xlsx"
    temp = nested_dir / "~$PCR_RUN.xlsx"
    ignored = nested_dir / "notes.txt"

    flat.write_text("flat", encoding="utf-8")
    nested.write_text("nested", encoding="utf-8")
    temp.write_text("temp", encoding="utf-8")
    ignored.write_text("ignored", encoding="utf-8")

    paths = _expand_input_paths(
        (str(tmp_path),),
        recursive_dirs=True,
        allowed_extensions={".xlsx", ".xlsm", ".xltx", ".xltm"},
    )

    assert paths == [nested, flat]


def test_cli_ingest_text_creates_schema_and_rows(tmp_path) -> None:
    db_path = tmp_path / "fis.sqlite"
    source_path = tmp_path / "source.txt"
    source_path.write_text("a;b;c\n1;2;3\n", encoding="utf-8")

    result = CliRunner().invoke(
        cli,
        [
            "ingest-text",
            "--db-path",
            str(db_path),
            "--source-name",
            "TEST",
            "--entity",
            "LAB",
            str(source_path),
        ],
    )

    assert result.exit_code == 0, result.output
    payload = json.loads(result.stdout)
    assert payload[0]["ok"] is True
    assert payload[0]["inserted_lines"] == 2
    assert "Ingesting text files" in result.stderr
    assert "Text ingest complete: 1 succeeded" in result.stderr

    engine = create_engine(f"sqlite:///{db_path}", future=True)
    with engine.begin() as conn:
        count = conn.execute(text("SELECT COUNT(*) FROM raw_text_lines")).scalar_one()

    assert count == 2


def test_cli_ingest_excel_skips_empty_workbook(tmp_path) -> None:
    from openpyxl import Workbook

    db_path = tmp_path / "fis.sqlite"
    valid_path = tmp_path / "valid.xlsx"
    empty_path = tmp_path / "empty.xlsx"

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["a", "b"])
    worksheet.append([1, 2])
    workbook.save(valid_path)
    empty_path.write_bytes(b"")

    result = CliRunner().invoke(
        cli,
        [
            "ingest-excel",
            "--db-path",
            str(db_path),
            "--source-name",
            "TEST",
            "--entity",
            "LAB",
            str(tmp_path),
        ],
    )

    assert result.exit_code == 0
    payload = json.loads(result.stdout)
    assert [item["ok"] for item in payload] == [False, True]
    assert payload[0]["skipped"] is True
    assert "empty" in payload[0]["error"]
    assert "Ingesting Excel workbooks" in result.stderr
    assert "Excel ingest complete: 1 succeeded, 1 skipped, 0 failed" in result.stderr

    engine = create_engine(f"sqlite:///{db_path}", future=True)
    with engine.begin() as conn:
        count = conn.execute(text("SELECT COUNT(*) FROM raw_excel_rows")).scalar_one()

    assert count == 2


def test_cli_ingest_excel_strict_fails_after_empty_workbook(tmp_path) -> None:
    db_path = tmp_path / "fis.sqlite"
    empty_path = tmp_path / "empty.xlsx"
    empty_path.write_bytes(b"")

    result = CliRunner().invoke(
        cli,
        [
            "ingest-excel",
            "--strict",
            "--db-path",
            str(db_path),
            "--source-name",
            "TEST",
            "--entity",
            "LAB",
            str(empty_path),
        ],
    )

    assert result.exit_code == 1
    payload = json.loads(result.stdout)
    assert payload[0]["skipped"] is True
    assert "Excel ingest complete: 0 succeeded, 1 skipped, 0 failed" in result.stderr


def test_cli_parse_excel_marks_unsupported_entity(tmp_path) -> None:
    db_path = tmp_path / "fis.sqlite"
    source_path = tmp_path / "unknown.xlsx"
    _write_workbook(
        source_path,
        sheets={"Sheet1": [["sample", "value"], ["S1", 1]]},
    )

    ingest_result = CliRunner().invoke(
        cli,
        [
            "ingest-excel",
            "--db-path",
            str(db_path),
            "--source-name",
            "unknown",
            "--entity",
            "unknown",
            str(source_path),
        ],
    )

    assert ingest_result.exit_code == 0, ingest_result.output

    result = CliRunner().invoke(
        cli,
        [
            "parse-excel",
            "--db-path",
            str(db_path),
            "--entity",
            "unknown",
        ],
    )

    assert result.exit_code == 0, result.output
    payload = json.loads(result.stdout)
    assert payload[0]["ok"] is True
    assert payload[0]["status"] == "UNSUPPORTED_ENTITY"
    assert payload[0]["rows_seen"] == 2
    assert "Parsing Excel files" in result.stderr
    assert "Excel parse complete: 1 succeeded, 0 failed" in result.stderr

    engine = create_engine(f"sqlite:///{db_path}", future=True)
    with engine.begin() as conn:
        raw_statuses = conn.execute(
            text("SELECT DISTINCT parse_status FROM raw_excel_rows")
        ).scalars().all()
        file_status = conn.execute(
            text("SELECT status FROM ctl_excel_parse_file")
        ).scalar_one()
        sheet_status = conn.execute(
            text("SELECT status FROM ctl_excel_parse_sheet")
        ).scalar_one()

    assert raw_statuses == ["UNSUPPORTED_ENTITY"]
    assert file_status == "UNSUPPORTED_ENTITY"
    assert sheet_status == "UNSUPPORTED_ENTITY"


def test_cli_parse_excel_uses_farmacia_default_profile(tmp_path) -> None:
    db_path = tmp_path / "fis.sqlite"
    source_path = tmp_path / "farmacia.xlsx"
    _write_workbook(
        source_path,
        sheets={"Dispensaciones": [["drug", "qty"], ["A", 2]]},
    )

    ingest_result = CliRunner().invoke(
        cli,
        [
            "ingest-excel",
            "--db-path",
            str(db_path),
            "--source-name",
            "farmacia",
            "--entity",
            "farmacia",
            str(source_path),
        ],
    )
    assert ingest_result.exit_code == 0, ingest_result.output

    result = CliRunner().invoke(
        cli,
        [
            "parse-excel",
            "--db-path",
            str(db_path),
            "--entity",
            "farmacia",
        ],
    )

    assert result.exit_code == 0, result.output
    payload = json.loads(result.stdout)
    assert payload[0]["file_kind"] == "default"
    assert payload[0]["status"] == "SKIPPED_METADATA"
    assert payload[0]["rows_seen"] == 2

    engine = create_engine(f"sqlite:///{db_path}", future=True)
    with engine.begin() as conn:
        audit = conn.execute(
            text(
                """
                SELECT f.file_kind, s.sheet_kind, s.header_row_no, s.status
                FROM ctl_excel_parse_file f
                JOIN ctl_excel_parse_sheet s
                  ON s.excel_file_parse_id = f.excel_file_parse_id
                """
            )
        ).one()

    assert audit == ("default", "default", 1, "SKIPPED_METADATA")


def test_cli_parse_excel_pcr_classifies_file_kind_from_filename(tmp_path) -> None:
    db_path = tmp_path / "fis.sqlite"
    source_path = tmp_path / "PCR_RUN - Quantification Cq Results.xlsx"
    _write_workbook(
        source_path,
        sheets={
            "0": [
                [
                    "Well",
                    "Fluor",
                    "Target",
                    "Content",
                    "Sample",
                    "Cq",
                    "Cq Mean",
                    "Cq Std. Dev.",
                ],
                ["A01", "FAM", "N1", "Unkn", "S1", 23.4, 23.4, 0.1],
            ],
            "Run Information": [["Run Name", "PCR_RUN"]],
        },
    )

    ingest_result = CliRunner().invoke(
        cli,
        [
            "ingest-excel",
            "--db-path",
            str(db_path),
            "--source-name",
            "pcr",
            "--entity",
            "pcr",
            str(source_path),
        ],
    )
    assert ingest_result.exit_code == 0, ingest_result.output

    result = CliRunner().invoke(
        cli,
        [
            "parse-excel",
            "--db-path",
            str(db_path),
            "--entity",
            "pcr",
        ],
    )

    assert result.exit_code == 0, result.output
    payload = json.loads(result.stdout)
    assert payload[0]["file_kind"] == "quantification_cq_results"

    engine = create_engine(f"sqlite:///{db_path}", future=True)
    with engine.begin() as conn:
        file_kind = conn.execute(
            text("SELECT file_kind FROM ctl_excel_parse_file")
        ).scalar_one()

    assert file_kind == "quantification_cq_results"


def test_cli_parse_excel_pcr_entity_matching_is_case_insensitive(tmp_path) -> None:
    db_path = tmp_path / "fis.sqlite"
    source_path = tmp_path / "PCR_RUN - Quantification Cq Results.xlsx"
    _write_workbook(
        source_path,
        sheets={
            "0": [
                [
                    "Well",
                    "Fluor",
                    "Target",
                    "Content",
                    "Sample",
                    "Cq",
                    "Cq Mean",
                    "Cq Std. Dev.",
                ],
                ["A01", "FAM", "N1", "Unkn", "S1", 23.4, 23.4, 0.1],
            ],
        },
    )

    ingest_result = CliRunner().invoke(
        cli,
        [
            "ingest-excel",
            "--db-path",
            str(db_path),
            "--source-name",
            "PCR",
            "--entity",
            "PCR",
            str(source_path),
        ],
    )
    assert ingest_result.exit_code == 0, ingest_result.output

    result = CliRunner().invoke(
        cli,
        [
            "parse-excel",
            "--db-path",
            str(db_path),
            "--entity",
            "PCR",
        ],
    )

    assert result.exit_code == 0, result.output
    payload = json.loads(result.stdout)
    assert payload[0]["entity_name"] == "PCR"
    assert payload[0]["file_kind"] == "quantification_cq_results"
    assert payload[0]["parser_name"] == "pcr-excel"
    assert payload[0]["status"] == "PARSED_OK"


def test_cli_parse_excel_pcr_parses_quantification_cq_results(tmp_path) -> None:
    db_path = tmp_path / "fis.sqlite"
    source_path = tmp_path / "PCR_RUN - Quantification Cq Results.xlsx"
    _write_workbook(
        source_path,
        sheets={
            "0": [
                ["Instrument", "redacted"],
                [
                    "Well",
                    "Fluor",
                    "Target",
                    "Content",
                    "Sample",
                    "Cq",
                    "Cq Mean",
                    "Cq Std. Dev.",
                ],
                ["A01", "FAM", "N1", "Unkn", "S1", 23.4, 23.5, 0.1],
                ["A02", "FAM", "N2", "Unkn", "S2", "not-a-number", 0, 0],
                [None, None, None, None],
            ],
            "Run Information": [["Run Name", "PCR_RUN"], ["Instrument", "QS7"]],
        },
    )

    ingest_result = CliRunner().invoke(
        cli,
        [
            "ingest-excel",
            "--db-path",
            str(db_path),
            "--source-name",
            "pcr",
            "--entity",
            "pcr",
            str(source_path),
        ],
    )
    assert ingest_result.exit_code == 0, ingest_result.output

    result = CliRunner().invoke(
        cli,
        [
            "parse-excel",
            "--db-path",
            str(db_path),
            "--entity",
            "pcr",
        ],
    )

    assert result.exit_code == 0, result.output
    payload = json.loads(result.stdout)
    assert payload[0]["file_kind"] == "quantification_cq_results"
    assert payload[0]["status"] == "PARSE_ERROR"
    assert payload[0]["rows_seen"] == 6
    assert payload[0]["rows_parsed"] == 3
    assert payload[0]["rows_error"] == 1

    engine = create_engine(f"sqlite:///{db_path}", future=True)
    with engine.begin() as conn:
        rows = conn.execute(
            text(
                """
                SELECT row_no, parse_status, payload_json, parse_error
                FROM raw_excel_rows
                WHERE sheet_name = '0'
                ORDER BY row_no ASC
                """
            )
        ).all()
        sheet_audit = conn.execute(
            text(
                """
                SELECT sheet_kind, header_row_no, status, rows_parsed, rows_error
                FROM ctl_excel_parse_sheet
                WHERE sheet_name = '0'
                """
            )
        ).one()
        run_info_payloads = conn.execute(
            text(
                """
                SELECT payload_json
                FROM raw_excel_rows
                WHERE sheet_name = 'Run Information'
                ORDER BY row_no ASC
                """
            )
        ).scalars().all()

    assert rows[0][1] == "SKIPPED_METADATA"
    assert rows[1][1] == "SKIPPED_METADATA"
    assert rows[2][1] == "PARSED_OK"
    assert rows[3][1] == "PARSE_ERROR"
    parsed_payload = json.loads(rows[2][2])
    assert parsed_payload == {
        "entity": "pcr",
        "file_kind": "quantification_cq_results",
        "sheet_kind": "cq_results",
        "fields": {
            "well": "A01",
            "fluor": "FAM",
            "target": "N1",
            "content": "Unkn",
            "sample": "S1",
            "cq": 23.4,
            "cq_mean": 23.5,
            "cq_std_dev": 0.1,
        },
        "source": {
            "file_id": 1,
            "sheet_name": "0",
            "row_no": 3,
            "header_row_no": 2,
        },
    }
    assert "Invalid numeric value for cq" in rows[3][3]
    assert sheet_audit == ("cq_results", 2, "PARSE_ERROR", 1, 1)
    assert json.loads(run_info_payloads[0])["fields"] == {
        "key": "Run Name",
        "value": "PCR_RUN",
    }


def test_cli_parse_excel_pcr_parses_quantification_amplification_results(
    tmp_path,
) -> None:
    db_path = tmp_path / "fis.sqlite"
    source_path = tmp_path / "PCR_RUN - Quantification Amplification Results.xlsx"
    _write_workbook(
        source_path,
        sheets={
            "SYBR": [
                [None, "Cycle", "A1", "A2"],
                [None, 1, 10.5, 20],
                [None, 2, None, 30.25],
                [None, 3, "bad-rfu", 40],
            ],
            "Run Information": [["Run Name", "PCR_RUN"], ["Instrument", "QS7"]],
        },
    )

    ingest_result = CliRunner().invoke(
        cli,
        [
            "ingest-excel",
            "--db-path",
            str(db_path),
            "--source-name",
            "PCR",
            "--entity",
            "PCR",
            str(source_path),
        ],
    )
    assert ingest_result.exit_code == 0, ingest_result.output

    result = CliRunner().invoke(
        cli,
        [
            "parse-excel",
            "--db-path",
            str(db_path),
            "--entity",
            "PCR",
        ],
    )

    assert result.exit_code == 0, result.output
    payload = json.loads(result.stdout)
    assert payload[0]["file_kind"] == "quantification_amplification_results"
    assert payload[0]["status"] == "PARSE_ERROR"
    assert payload[0]["rows_seen"] == 6
    assert payload[0]["rows_parsed"] == 4
    assert payload[0]["rows_error"] == 1

    engine = create_engine(f"sqlite:///{db_path}", future=True)
    with engine.begin() as conn:
        rows = conn.execute(
            text(
                """
                SELECT row_no, parse_status, payload_json, parse_error
                FROM raw_excel_rows
                WHERE sheet_name = 'SYBR'
                ORDER BY row_no ASC
                """
            )
        ).all()
        sheet_audit = conn.execute(
            text(
                """
                SELECT sheet_kind, header_row_no, status, rows_parsed, rows_error
                FROM ctl_excel_parse_sheet
                WHERE sheet_name = 'SYBR'
                """
            )
        ).one()

    assert rows[0][1] == "SKIPPED_METADATA"
    assert rows[1][1] == "PARSED_OK"
    assert rows[2][1] == "PARSED_OK"
    assert rows[3][1] == "PARSE_ERROR"
    first_payload = json.loads(rows[1][2])
    assert first_payload == {
        "entity": "pcr",
        "file_kind": "quantification_amplification_results",
        "sheet_kind": "amplification_sybr",
        "fields": {
            "cycle": 1,
            "rfu_by_well": {
                "A1": 10.5,
                "A2": 20.0,
            },
        },
        "source": {
            "file_id": 1,
            "sheet_name": "SYBR",
            "row_no": 2,
            "header_row_no": 1,
        },
    }
    second_payload = json.loads(rows[2][2])
    assert second_payload["fields"]["rfu_by_well"] == {"A1": None, "A2": 30.25}
    assert "Invalid RFU value for well A1" in rows[3][3]
    assert sheet_audit == ("amplification_sybr", 1, "PARSE_ERROR", 2, 1)


def test_cli_parse_excel_pcr_parses_melt_curve_rfu_results(tmp_path) -> None:
    db_path = tmp_path / "fis.sqlite"
    source_path = tmp_path / "PCR_RUN - Melt Curve RFU Results.xlsx"
    _write_workbook(
        source_path,
        sheets={
            "SYBR": [
                [None, "Temperature", "A1", "A2"],
                [None, 65.0, 100, 200],
                [None, 65.5, "", 210.25],
                [None, "bad-temp", 120, 220],
            ],
            "Run Information": [["Run Name", "PCR_RUN"], ["Instrument", "QS7"]],
        },
    )

    ingest_result = CliRunner().invoke(
        cli,
        [
            "ingest-excel",
            "--db-path",
            str(db_path),
            "--source-name",
            "PCR",
            "--entity",
            "PCR",
            str(source_path),
        ],
    )
    assert ingest_result.exit_code == 0, ingest_result.output

    result = CliRunner().invoke(
        cli,
        [
            "parse-excel",
            "--db-path",
            str(db_path),
            "--entity",
            "PCR",
        ],
    )

    assert result.exit_code == 0, result.output
    payload = json.loads(result.stdout)
    assert payload[0]["file_kind"] == "melt_curve_rfu_results"
    assert payload[0]["status"] == "PARSE_ERROR"
    assert payload[0]["rows_seen"] == 6
    assert payload[0]["rows_parsed"] == 4
    assert payload[0]["rows_error"] == 1

    engine = create_engine(f"sqlite:///{db_path}", future=True)
    with engine.begin() as conn:
        rows = conn.execute(
            text(
                """
                SELECT row_no, parse_status, payload_json, parse_error
                FROM raw_excel_rows
                WHERE sheet_name = 'SYBR'
                ORDER BY row_no ASC
                """
            )
        ).all()
        sheet_audit = conn.execute(
            text(
                """
                SELECT sheet_kind, header_row_no, status, rows_parsed, rows_error
                FROM ctl_excel_parse_sheet
                WHERE sheet_name = 'SYBR'
                """
            )
        ).one()

    assert rows[0][1] == "SKIPPED_METADATA"
    assert rows[1][1] == "PARSED_OK"
    assert rows[2][1] == "PARSED_OK"
    assert rows[3][1] == "PARSE_ERROR"
    first_payload = json.loads(rows[1][2])
    assert first_payload == {
        "entity": "pcr",
        "file_kind": "melt_curve_rfu_results",
        "sheet_kind": "melt_curve_sybr",
        "fields": {
            "temperature": 65.0,
            "rfu_by_well": {
                "A1": 100.0,
                "A2": 200.0,
            },
        },
        "source": {
            "file_id": 1,
            "sheet_name": "SYBR",
            "row_no": 2,
            "header_row_no": 1,
        },
    }
    second_payload = json.loads(rows[2][2])
    assert second_payload["fields"]["rfu_by_well"] == {"A1": None, "A2": 210.25}
    assert "Invalid numeric value for temperature" in rows[3][3]
    assert sheet_audit == ("melt_curve_sybr", 1, "PARSE_ERROR", 2, 1)


def test_cli_parse_excel_pcr_parses_melt_curve_derivative_results(
    tmp_path,
) -> None:
    db_path = tmp_path / "fis.sqlite"
    source_path = tmp_path / "PCR_RUN - Melt Curve Derivative Results.xlsx"
    _write_workbook(
        source_path,
        sheets={
            "SYBR": [
                [None, "Temperature", "A1", "A2"],
                [None, 65.0, 100, 200],
            ],
            "Run Information": [["Run Name", "PCR_RUN"]],
        },
    )

    ingest_result = CliRunner().invoke(
        cli,
        [
            "ingest-excel",
            "--db-path",
            str(db_path),
            "--source-name",
            "PCR",
            "--entity",
            "PCR",
            str(source_path),
        ],
    )
    assert ingest_result.exit_code == 0, ingest_result.output

    result = CliRunner().invoke(
        cli,
        [
            "parse-excel",
            "--db-path",
            str(db_path),
            "--entity",
            "PCR",
        ],
    )

    assert result.exit_code == 0, result.output
    payload = json.loads(result.stdout)
    assert payload[0]["file_kind"] == "melt_curve_derivative_results"
    assert payload[0]["status"] == "PARSED_OK"
    assert payload[0]["rows_seen"] == 3
    assert payload[0]["rows_parsed"] == 2

    engine = create_engine(f"sqlite:///{db_path}", future=True)
    with engine.begin() as conn:
        parsed_payload = conn.execute(
            text(
                """
                SELECT payload_json
                FROM raw_excel_rows
                WHERE sheet_name = 'SYBR' AND parse_status = 'PARSED_OK'
                """
            )
        ).scalar_one()
        sheet_audit = conn.execute(
            text(
                """
                SELECT sheet_kind, header_row_no, status, rows_parsed, rows_error
                FROM ctl_excel_parse_sheet
                WHERE sheet_name = 'SYBR'
                """
            )
        ).one()

    assert json.loads(parsed_payload) == {
        "entity": "pcr",
        "file_kind": "melt_curve_derivative_results",
        "sheet_kind": "melt_curve_derivative_sybr",
        "fields": {
            "temperature": 65.0,
            "rfu_by_well": {
                "A1": 100.0,
                "A2": 200.0,
            },
        },
        "source": {
            "file_id": 1,
            "sheet_name": "SYBR",
            "row_no": 2,
            "header_row_no": 1,
        },
    }
    assert sheet_audit == ("melt_curve_derivative_sybr", 1, "PARSED_OK", 1, 0)


def test_cli_parse_excel_pcr_missing_required_header_is_schema_mismatch(
    tmp_path,
) -> None:
    db_path = tmp_path / "fis.sqlite"
    source_path = tmp_path / "PCR_RUN - Quantification Cq Results.xlsx"
    _write_workbook(
        source_path,
        sheets={"0": [["Well", "Sample Name", "Cq"], ["A01", "S1", 23.4]]},
    )

    ingest_result = CliRunner().invoke(
        cli,
        [
            "ingest-excel",
            "--db-path",
            str(db_path),
            "--source-name",
            "pcr",
            "--entity",
            "pcr",
            str(source_path),
        ],
    )
    assert ingest_result.exit_code == 0, ingest_result.output

    result = CliRunner().invoke(
        cli,
        [
            "parse-excel",
            "--db-path",
            str(db_path),
            "--entity",
            "pcr",
        ],
    )

    assert result.exit_code == 0, result.output
    payload = json.loads(result.stdout)
    assert payload[0]["status"] == "SCHEMA_MISMATCH"

    engine = create_engine(f"sqlite:///{db_path}", future=True)
    with engine.begin() as conn:
        statuses = conn.execute(
            text("SELECT DISTINCT parse_status FROM raw_excel_rows")
        ).scalars().all()
        error = conn.execute(
            text("SELECT error FROM ctl_excel_parse_sheet")
        ).scalar_one()

    assert statuses == ["SCHEMA_MISMATCH"]
    assert error == "No matching header row found."


def test_cli_parse_excel_pcr_unknown_file_and_sheet_are_explicit_skips(
    tmp_path,
) -> None:
    db_path = tmp_path / "fis.sqlite"
    unknown_file = tmp_path / "PCR_RUN - Unknown Export.xlsx"
    unknown_sheet_file = tmp_path / "PCR_RUN - Quantification Cq Results.xlsx"
    _write_workbook(
        unknown_file,
        sheets={
            "Results": [
                ["Well", "Sample Name", "Target Name", "Cq"],
                ["A01", "S1", "N1", 1],
            ]
        },
    )
    _write_workbook(
        unknown_sheet_file,
        sheets={
            "Notes": [
                ["Well", "Sample Name", "Target Name", "Cq"],
                ["A01", "S1", "N1", 1],
            ]
        },
    )

    ingest_result = CliRunner().invoke(
        cli,
        [
            "ingest-excel",
            "--db-path",
            str(db_path),
            "--source-name",
            "pcr",
            "--entity",
            "pcr",
            str(tmp_path),
        ],
    )
    assert ingest_result.exit_code == 0, ingest_result.output

    result = CliRunner().invoke(
        cli,
        [
            "parse-excel",
            "--db-path",
            str(db_path),
            "--entity",
            "pcr",
        ],
    )

    assert result.exit_code == 0, result.output
    payload = sorted(
        json.loads(result.stdout),
        key=lambda item: item["file_kind"] or "",
    )
    assert [item["status"] for item in payload] == [
        "SKIPPED_FILE_PROFILE",
        "SKIPPED_SHEET_PROFILE",
    ]

    engine = create_engine(f"sqlite:///{db_path}", future=True)
    with engine.begin() as conn:
        statuses = conn.execute(
            text(
                """
                SELECT DISTINCT parse_status
                FROM raw_excel_rows
                ORDER BY parse_status ASC
                """
            )
        ).scalars().all()

    assert statuses == ["SKIPPED_FILE_PROFILE", "SKIPPED_SHEET_PROFILE"]


def test_cli_parse_excel_filters_and_reprocesses(tmp_path) -> None:
    db_path = tmp_path / "fis.sqlite"
    first_path = tmp_path / "first.xlsx"
    second_path = tmp_path / "second.xlsx"
    _write_workbook(
        first_path,
        sheets={
            "Target": [["sample", "value"], ["S1", 1]],
            "Other": [["sample", "value"], ["S2", 2]],
        },
    )
    _write_workbook(
        second_path,
        sheets={"Target": [["sample", "value"], ["S3", 3]]},
    )

    ingest_result = CliRunner().invoke(
        cli,
        [
            "ingest-excel",
            "--db-path",
            str(db_path),
            "--source-name",
            "farmacia",
            "--entity",
            "farmacia",
            str(tmp_path),
        ],
    )
    assert ingest_result.exit_code == 0, ingest_result.output
    ingested = json.loads(ingest_result.stdout)
    first_file_id = next(
        item["file_id"] for item in ingested if item["path"] == str(first_path)
    )

    result = CliRunner().invoke(
        cli,
        [
            "parse-excel",
            "--db-path",
            str(db_path),
            "--entity",
            "farmacia",
            "--file-id",
            str(first_file_id),
            "--sheet-name",
            "Target",
        ],
    )

    assert result.exit_code == 0, result.output
    payload = json.loads(result.stdout)
    assert [item["file_id"] for item in payload] == [first_file_id]
    assert payload[0]["rows_seen"] == 2

    no_pending_result = CliRunner().invoke(
        cli,
        [
            "parse-excel",
            "--db-path",
            str(db_path),
            "--entity",
            "farmacia",
            "--file-id",
            str(first_file_id),
            "--sheet-name",
            "Target",
        ],
    )
    assert no_pending_result.exit_code == 0, no_pending_result.output
    assert json.loads(no_pending_result.stdout) == []

    reprocess_result = CliRunner().invoke(
        cli,
        [
            "parse-excel",
            "--reprocess-all",
            "--db-path",
            str(db_path),
            "--entity",
            "farmacia",
            "--file-id",
            str(first_file_id),
            "--sheet-name",
            "Target",
        ],
    )

    assert reprocess_result.exit_code == 0, reprocess_result.output
    reprocess_payload = json.loads(reprocess_result.stdout)
    assert [item["file_id"] for item in reprocess_payload] == [first_file_id]
    assert reprocess_payload[0]["rows_seen"] == 2

    engine = create_engine(f"sqlite:///{db_path}", future=True)
    with engine.begin() as conn:
        parsed_sheets = conn.execute(
            text(
                """
                SELECT sheet_name
                FROM ctl_excel_parse_sheet
                WHERE file_id = :file_id
                ORDER BY excel_sheet_parse_id ASC
                """
            ),
            {"file_id": first_file_id},
        ).scalars().all()
        second_file_audits = conn.execute(
            text(
                """
                SELECT COUNT(*)
                FROM ctl_excel_parse_file
                WHERE file_id != :file_id
                """
            ),
            {"file_id": first_file_id},
        ).scalar_one()

    assert parsed_sheets == ["Target", "Target"]
    assert second_file_audits == 0


def test_cli_parse_microb_shows_progress(tmp_path) -> None:
    db_path = tmp_path / "fis.sqlite"
    source_path = tmp_path / "microb.txt"
    source_path.write_text("a|b|c|\n1|v2|v3|\n", encoding="utf-8")

    ingest_result = CliRunner().invoke(
        cli,
        [
            "ingest-text",
            "--db-path",
            str(db_path),
            "--source-name",
            "microb",
            "--entity",
            "microb",
            str(source_path),
        ],
    )

    assert ingest_result.exit_code == 0, ingest_result.output

    result = CliRunner().invoke(
        cli,
        [
            "parse-microb",
            "--db-path",
            str(db_path),
            "--entity",
            "microb",
        ],
    )

    assert result.exit_code == 0, result.output
    payload = json.loads(result.stdout)
    assert payload[0]["ok"] is True
    assert payload[0]["parsed_total"] == 1
    assert payload[0]["parsed_ok"] == 1
    assert "Parsing Microb files" in result.stderr
    assert "Microb parse complete: 1 succeeded, 0 failed" in result.stderr

    no_pending_result = CliRunner().invoke(
        cli,
        [
            "parse-microb",
            "--db-path",
            str(db_path),
            "--entity",
            "microb",
        ],
    )

    assert no_pending_result.exit_code == 0, no_pending_result.output
    assert json.loads(no_pending_result.stdout) == []

    reprocess_result = CliRunner().invoke(
        cli,
        [
            "parse-microb",
            "--reprocess-all",
            "--db-path",
            str(db_path),
            "--entity",
            "microb",
        ],
    )

    assert reprocess_result.exit_code == 0, reprocess_result.output
    reprocess_payload = json.loads(reprocess_result.stdout)
    assert reprocess_payload[0]["ok"] is True
    assert reprocess_payload[0]["parsed_total"] == 1
    assert reprocess_payload[0]["parsed_ok"] == 1
    assert "Parsing Microb files" in reprocess_result.stderr
    assert "Microb parse complete: 1 succeeded, 0 failed" in reprocess_result.stderr
