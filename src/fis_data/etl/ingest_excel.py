"""Raw ingestion for Excel files."""

from __future__ import annotations

import hashlib
import json
from pathlib import Path
from typing import Any
from zipfile import BadZipFile

from sqlalchemy import text
from sqlalchemy.engine import Engine


def row_sha256(values: list[Any]) -> str:
    """Compute a stable SHA256 digest for an Excel row."""

    payload = json.dumps(values, ensure_ascii=False, default=str, separators=(",", ":"))
    return hashlib.sha256(payload.encode("utf-8", errors="replace")).hexdigest()


def ingest_excel_file(
    engine: Engine,
    *,
    source_id: int,
    entity_name: str,
    run_id: int,
    file_id: int,
    path: str | Path,
    batch_size: int = 1000,
) -> dict[str, int]:
    """Ingest workbook rows into ``raw_excel_rows``.

    ``openpyxl`` is imported lazily so text-only installations remain lightweight.
    """

    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "Excel ingestion requires openpyxl. Install development dependencies "
            'or add "openpyxl>=3.1,<4" to the runtime environment.'
        ) from exc

    input_path = Path(path)
    if input_path.stat().st_size == 0:
        raise ValueError(f"Input Excel file is empty: {input_path}")

    try:
        workbook = load_workbook(input_path, read_only=True, data_only=True)
    except BadZipFile as exc:
        raise ValueError(
            f"Input Excel file is not a valid OpenXML workbook: {input_path}"
        ) from exc
    rows: list[dict[str, object]] = []
    inserted = 0
    total = 0

    insert_stmt = text(
        """
        INSERT OR IGNORE INTO raw_excel_rows (
            source_id, run_id, file_id, entity_name, sheet_name, row_no,
            raw_values_json, row_sha256, is_header, parse_status
        )
        VALUES (
            :source_id, :run_id, :file_id, :entity_name, :sheet_name, :row_no,
            :raw_values_json, :row_sha256, :is_header, 'RAW_ONLY'
        )
        """
    )

    try:
        with engine.begin() as conn:
            for sheet in workbook.worksheets:
                rows_iter = sheet.iter_rows(values_only=True)
                for row_no, row in enumerate(rows_iter, start=1):
                    values = list(row)
                    if all(value is None for value in values):
                        continue

                    total += 1
                    rows.append(
                        {
                            "source_id": source_id,
                            "run_id": run_id,
                            "file_id": file_id,
                            "entity_name": entity_name,
                            "sheet_name": sheet.title,
                            "row_no": row_no,
                            "raw_values_json": json.dumps(
                                values,
                                ensure_ascii=False,
                                default=str,
                                separators=(",", ":"),
                            ),
                            "row_sha256": row_sha256(values),
                            "is_header": row_no == 1,
                        }
                    )

                    if len(rows) >= batch_size:
                        result = conn.execute(insert_stmt, rows)
                        inserted += result.rowcount or 0
                        rows.clear()

            if rows:
                result = conn.execute(insert_stmt, rows)
                inserted += result.rowcount or 0
    finally:
        workbook.close()

    return {"inserted_rows": inserted, "total_rows": total}
